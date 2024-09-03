import openpyxl
from tkinter import Tk
from tkinter.filedialog import askopenfilename

# GUI 윈도우 숨기기
Tk().withdraw()

# 파일 선택 창 열기
file_path = askopenfilename(title="엑셀 파일을 선택하세요", filetypes=[("Excel files", "*.xlsx *.xls")])

if not file_path:
    print("파일이 선택되지 않았습니다.")
else:
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active  # 기본 시트 선택

    # 병합된 셀 범위를 확인
    # 리스트로 복사하여 안전하게 순회
    merged_ranges = list(sheet.merged_cells.ranges)

    for merged_range in merged_ranges:
        # 범위가 B열에 있는 경우에만 실행
        if merged_range.min_col == merged_range.max_col == 2:  # B열 (1-indexed)
            # 병합된 셀의 값을 가져옴
            merge_cell_value = sheet.cell(row=merged_range.min_row, column=2).value

            # 병합 해제
            sheet.unmerge_cells(str(merged_range))

            # 각 셀에 병합된 셀 값 입력
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                sheet.cell(row=row, column=2, value=merge_cell_value)

    # 변경된 내용을 엑셀 파일에 저장
    workbook.save(file_path)

    print("작업이 완료되었습니다. 병합된 셀의 데이터가 각 셀에 붙여넣어졌습니다.")
