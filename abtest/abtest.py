import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def select_file(title):
    """파일 선택 다이얼로그"""
    root = tk.Tk()
    root.withdraw()  # 루트 윈도우 숨기기
    file_path = filedialog.askopenfilename(title=title)
    return file_path

def apply_differences_to_excel(df_a, df_b, output_file):
    """비교 결과를 엑셀 파일에 색으로 표시하며 저장"""
    wb = Workbook()
    ws = wb.active

    # 헤더 추가
    for col_num, column_title in enumerate(df_a.columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)

    # 셀 색상 스타일
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row_index, row in df_a.iterrows():
        for col_index, value in enumerate(row, 1):
            a_value = df_a.iat[row_index, col_index - 1]
            b_value = df_b.iat[row_index, col_index - 1]

            cell = ws.cell(row=row_index + 2, column=col_index, value=a_value)

            if pd.isna(a_value) and pd.isna(b_value):
                continue

            if a_value != b_value:
                cell.fill = fill

    wb.save(output_file)

# Step 1: File Selection
file_a = select_file("A 엑셀 파일을 선택하세요.")
file_b = select_file("B 엑셀 파일을 선택하세요.")

# 에러 핸들링: 파일 선택이 정상적으로 완료되었는지 확인
if not file_a or not file_b:
    print("두 파일 중 하나 또는 모두를 선택하지 않았습니다. 다시 시도해주세요.")
else:
    # Step 2: Load Excel files into DataFrames
    df_a = pd.read_excel(file_a)
    df_b = pd.read_excel(file_b)

    # 인덱스와 컬럼 레이블을 동기화
    df_b = df_b.reindex(index=df_a.index, columns=df_a.columns)

    # Step 3: Apply differences to Excel with color
    output_file = 'C_비교_결과.xlsx'
    apply_differences_to_excel(df_a, df_b, output_file)
    print(f"비교 결과가 {output_file} 파일로 저장되었습니다.")
