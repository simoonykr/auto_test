import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from multiprocessing import Pool

def compare_chunk(param):
    start, end, df_a_chunk, df_b_chunk = param
    differences = []

    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row_index in range(len(df_a_chunk)):
        for col_index in range(len(df_a_chunk.columns)):
            a_value = df_a_chunk.iat[row_index, col_index]
            b_value = df_b_chunk.iat[row_index, col_index]

            if pd.isna(a_value) and pd.isna(b_value):
                continue

            if a_value != b_value:
                differences.append((start + row_index + 2, col_index + 1, a_value, fill))

    return differences

def apply_differences_to_excel_mp(df_a, df_b, output_file, num_chunks=4):
    chunksize = len(df_a) // num_chunks
    params = []
    for i in range(num_chunks):
        start = i * chunksize
        end = (i + 1) * chunksize if i != num_chunks - 1 else len(df_a)
        params.append((start, end, df_a.iloc[start:end], df_b.iloc[start:end]))

    with Pool(processes=num_chunks) as pool:
        results = pool.map(compare_chunk, params)

    # Flatten the list of results
    differences = [item for sublist in results for item in sublist]

    # Create a new workbook and apply differences
    wb = Workbook()
    ws = wb.active

    # Write headers
    for col_num, column_title in enumerate(df_a.columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)

    # Write data and apply styles
    for row_index, col_index, value, fill in differences:
        cell = ws.cell(row=row_index, column=col_index, value=value)
        cell.fill = fill

    wb.save(output_file)

# Load your data
df_a = pd.read_excel("")
df_b = pd.read_excel("")

# Ensure both dataframes have the same indexes and columns
df_b = df_b.reindex(index=df_a.index, columns=df_a.columns)

# Apply differences using your improved function
output_file = 'C_비교_결과.xlsx'
apply_differences_to_excel_mp(df_a, df_b, output_file)
print(f"Comparative results saved to {output_file}.")